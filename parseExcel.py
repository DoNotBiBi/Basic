import openpyxl

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font


def test_parseExcel():
    wb = openpyxl.load_workbook('a.xlsx')
    # print(wb.get_sheet_names()) # 输出所有工作表 list
    sheet = wb.get_sheet_by_name('Sheet1')  # 指定某个工作表
    highest_rows = sheet.max_row
    highest_columns = sheet.max_column
    print(highest_rows)
    print(highest_columns)
    for i in range(1, highest_rows + 1):
        print(str(sheet[f'A{i}'].value) + '    ' + str(sheet[f'B{i}'].value) + '    ' + str(sheet[f'C{i}'].value))

    for i in range(1, highest_rows + 1):
        print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value)

    print(get_column_letter(sheet.max_column))  # 获取最大列对应的字母
    print(column_index_from_string('AA'))  # 获取字母对应的列数

    print(tuple(sheet['A1':'C6']))
    for rowOfCellObjects in sheet['A1':'C6']:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)  # 获取对应坐标以及相关值
        print('------END OF ROW--------')

    print(sheet.columns)
    # 有意思的情况
    for cellObj2 in sheet.rows:  # 获取某一列的所有值
        print(cellObj2[1].value)

    # 创建工作表 若是没有传参数，则自动往后加
    # wb.create_sheet(index=0, title='test')
    print(wb.get_sheet_names())
    # 删除工作表
    # wb.remove_sheet(wb.get_sheet_by_name('xxx'))

    # 写入对应表
    # test = wb.get_sheet_by_name('test')
    # test['A1'] = 'hello world'
    # wb.save('a.xlsx')  # 千万记得保存

    # 设置字体
    italic24Font = Font(size=24, italic=True)
    sheet['A1'].font = italic24Font
    wb.save('a.xlsx')
    # 对于整个A列进行格式设置
    # sheet.column_dimensions['A'].font = italic24Font

    # todo:
    #  公式求和 有点小问题
    sheet['C7'] = '=SUM(C2:C6)'
    wb.save('a.xlsx')
    print(sheet['C7'].value)  # 若只是显示数据 则在wb = openpyxl.load_workbook('a.xlsx') 参数添加上data_only=True

    # 调整行高和列宽
    sheet.row_dimensions[1].height = 70
    sheet.column_dimensions['B'].width = 20
    wb.save('a.xlsx')

    # 合并单元格
    sheet.merge_cells('A10:C10')
    sheet['A10'] = 'hello wangkang'
    wb.save('a.xlsx')

    # 冻结某一行或者某一列 A1 或者 None 表示解冻
    sheet.freeze_panes = 'A2'
    wb.save('a.xlsx')

    # 制作图标
    refObj = openpyxl.chart.Reference(sheet, min_row = 2, min_col = 1, max_row = 6, max_col = 3)
    seriesObj=openpyxl.chart.Series(refObj,title='first series')
    chartObj=openpyxl.chart.BarChart()
    chartObj.append(seriesObj)
    chartObj.title='my chart'
    sheet.add_chart(chartObj,'c11')
    wb.save('a.xlsx')
